"""
Usage: python3 export_excel.py --data '[{"nom":"...","adresse":"...","cp":"...","ville":"...","tel":"...","site":"..."}]' --output /path/to/file.xlsx
"""
import argparse, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def export(data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Google Maps"

    headers = ["Nom", "Adresse", "Code Postal", "Ville", "Téléphone", "Site web"]
    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    for row, r in enumerate(data, 2):
        ws.cell(row, 1, value=r.get("nom", ""))
        ws.cell(row, 2, value=r.get("adresse", ""))
        ws.cell(row, 3, value=r.get("cp", ""))
        ws.cell(row, 4, value=r.get("ville", ""))
        ws.cell(row, 5, value=r.get("tel", ""))
        site = r.get("site", "")
        cell = ws.cell(row, 6, value=site)
        if site:
            cell.font = Font(color="0563C1", underline="single")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)

    wb.save(output_path)
    print(f"Saved: {output_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--data", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    export(json.loads(args.data), args.output)
